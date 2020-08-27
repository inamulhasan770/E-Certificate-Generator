import openpyxl
from PIL import Image
from PIL import ImageFont
from PIL import ImageDraw 
path = "data.xlsx"
wb_obj = openpyxl.load_workbook(path)   
sheet_obj = wb_obj.active 
m_row = sheet_obj.max_row
m_col = sheet_obj.max_column
for i in range(1, m_row+1):
    name = sheet_obj.cell(row = i+1, column = 2)
    email = sheet_obj.cell(row = i+1, column = 3)
    print(name.value, email.value)
    x = name.value;
    y = email.value;
    img = Image.open("cert.png")
    draw = ImageDraw.Draw(img)
    font = ImageFont.truetype("ttrb.ttf", 96)
    draw.text((1480,740), x, (0,0,0), font = font)
    flnm = y+".png"
    img.save(flnm)

